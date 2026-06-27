import pandas as pd
import openpyxl
from datetime import datetime
import os
import chardet

# 文件路径
source_file = r'E:\B01Python\PythonProject\A06_xop\xop_gld_history.csv'
target_file = r'E:\2026年QDII LOF基金估值偏差记录.xlsm'

# 与源程序保持一致的编码设置
CSV_ENCODING = 'utf-8-sig'  # 统一使用UTF-8 with BOM


def load_csv_with_encoding(csv_path):
    """
    智能加载CSV文件，与源程序保持一致的编码处理方式
    复制自您的XOP.py程序
    """
    if not os.path.exists(csv_path):
        return pd.DataFrame()

    # 检测文件编码
    try:
        with open(csv_path, 'rb') as f:
            raw_data = f.read()
            result = chardet.detect(raw_data)
            detected_encoding = result['encoding']
            print(f"  检测到文件编码: {detected_encoding} (置信度: {result['confidence']:.2f})")
    except Exception as e:
        print(f"  编码检测失败: {e}")
        detected_encoding = None

    # 尝试检测到的编码
    if detected_encoding:
        try:
            df = pd.read_csv(csv_path, encoding=detected_encoding)
            print(f"  ✓ 使用检测到的编码 {detected_encoding} 读取成功")
            return df
        except Exception as e:
            print(f"  使用检测到的编码失败: {e}")

    # 尝试常见编码列表（与源程序一致）
    encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'latin-1', 'cp1252']

    for encoding in encodings:
        try:
            df = pd.read_csv(csv_path, encoding=encoding)
            print(f"  ✓ 使用 {encoding} 编码读取成功")
            return df
        except UnicodeDecodeError:
            continue
        except Exception as e:
            print(f"  使用 {encoding} 编码时出错: {e}")
            continue

    # 最后尝试使用errors='ignore'
    try:
        df = pd.read_csv(csv_path, encoding='utf-8', errors='ignore')
        print(f"  ⚠ 使用忽略错误模式读取成功")
        return df
    except Exception as e:
        print(f"  ✗ 无法读取文件: {e}")
        return pd.DataFrame()


def update_xop_price():
    """
    从CSV源文件读取XOP数据，更新到Excel目标文件的XOP_price列
    """
    try:
        # 1. 使用与源程序一致的方式读取CSV文件
        print(f"正在读取源文件: {source_file}")
        df_source = load_csv_with_encoding(source_file)

        if df_source.empty:
            print("错误：无法读取CSV文件或文件为空")
            return

        # 显示所有列名，方便确认
        print(f"\n源文件列名: {df_source.columns.tolist()}")
        print(f"源文件行数: {len(df_source)}")

        # 检查是否有XOP列
        if 'XOP' not in df_source.columns:
            print("错误：源文件中没有找到'XOP'列")
            print(f"可用的列: {df_source.columns.tolist()}")
            return

        # 检查日期列
        date_column = None
        for col in df_source.columns:
            if col in ['日期', 'date', 'trade_date', '交易日期']:
                date_column = col
                break

        if date_column is None:
            # 如果没有找到明确的日期列名，使用第一列
            date_column = df_source.columns[0]
            print(f"使用第一列作为日期列: {date_column}")
        else:
            print(f"找到日期列: {date_column}")

        # 转换日期格式，确保统一
        df_source[date_column] = pd.to_datetime(df_source[date_column]).dt.date

        # 创建日期到XOP价格的映射字典
        price_dict = dict(zip(df_source[date_column], df_source['XOP']))
        print(f"共加载 {len(price_dict)} 条XOP价格数据")

        # 显示前几条数据供确认
        print("\n源数据预览（前5行）:")
        print(df_source[[date_column, 'XOP']].head())

        # 显示数据统计信息
        print(f"\n数据统计:")
        print(f"  日期范围: {df_source[date_column].min()} 至 {df_source[date_column].max()}")
        print(f"  价格范围: {df_source['XOP'].min():.2f} 至 {df_source['XOP'].max():.2f}")

        # 检查是否有空值
        null_count = df_source['XOP'].isnull().sum()
        if null_count > 0:
            print(f"  ⚠ XOP列有 {null_count} 个空值")

        # 2. 打开Excel目标文件
        print(f"\n正在打开目标文件: {target_file}")
        wb = openpyxl.load_workbook(target_file, keep_vba=True,data_only=True)

        # 获取活动工作表
        ws = wb.active
        print(f"当前工作表: {ws.title}")

        # 可选：指定特定工作表
        # ws = wb['Sheet1']  # 如果需要在特定工作表操作

        # 3. 查找日期列和XOP_price列
        header_row = 1
        date_col_idx = None
        xop_price_col_idx = None

        print("\n正在查找Excel表头...")
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value:
                cell_str = str(cell_value).strip().lower()
                print(f"  列{col_idx}: {cell_value}")

                if '日期' in cell_str or 'date' in cell_str:
                    date_col_idx = col_idx
                    print(f"  ✅ 找到日期列: 第{col_idx}列 (标题: {cell_value})")
                elif 'xop_price' in cell_str.replace(' ', '').lower():
                    xop_price_col_idx = col_idx
                    print(f"  ✅ 找到XOP_price列: 第{col_idx}列 (标题: {cell_value})")

        if date_col_idx is None:
            print("\n❌ 错误：未找到日期列")
            print("请确保表头包含'日期'或'date'关键字")
            wb.close()
            return

        if xop_price_col_idx is None:
            print("\n❌ 错误：未找到XOP_price列")
            print("请确保表头包含'XOP_price'关键字")
            wb.close()
            return

        # 4. 遍历数据行，匹配日期并更新XOP价格
        print(f"\n开始更新数据...")
        updated_count = 0
        not_found_count = 0
        date_not_parsed = 0
        skipped_empty = 0

        for row_idx in range(header_row + 1, ws.max_row + 1):
            date_cell = ws.cell(row=row_idx, column=date_col_idx)
            price_cell = ws.cell(row=row_idx, column=xop_price_col_idx)

            # 跳过空日期
            if date_cell.value is None:
                skipped_empty += 1
                continue

            # 转换日期格式
            try:
                if isinstance(date_cell.value, datetime):
                    current_date = date_cell.value.date()
                else:
                    # 尝试多种日期格式
                    date_str = str(date_cell.value).strip()
                    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y', '%Y%m%d']:
                        try:
                            current_date = datetime.strptime(date_str, fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        # 如果所有格式都失败，使用pandas解析
                        current_date = pd.to_datetime(date_str).date()
            except Exception as e:
                date_not_parsed += 1
                continue

            # 查找对应的XOP价格
            if current_date in price_dict:
                price_value = float(price_dict[current_date])
                price_cell.value = price_value
                updated_count += 1
                # 每100条显示一次进度
                if updated_count % 100 == 0:
                    print(f"  已更新 {updated_count} 条记录...")
            else:
                not_found_count += 1

        # 5. 保存文件
        print(f"\n正在保存文件: {target_file}")
        wb.save(target_file)
        wb.close()

        # 6. 输出统计结果
        print("\n" + "=" * 60)
        print(f"✅ 更新完成!")
        print(f"✅ 成功更新: {updated_count} 条记录")
        print(f"⚠️  未找到匹配数据: {not_found_count} 条记录")
        print(f"⚠️  日期解析失败: {date_not_parsed} 条记录")
        print(f"ℹ️  跳过空日期行: {skipped_empty} 条记录")
        print(f"📁 文件已保存至: {target_file}")
        print("=" * 60)

        # 显示最后几条更新示例
        if updated_count > 0:
            print("\n最后5条更新示例:")
            # 获取最后5个更新的日期和价格
            last_dates = list(price_dict.keys())[-5:]
            for date in last_dates:
                print(f"  {date}: {price_dict[date]:.2f}")

    except FileNotFoundError as e:
        print(f"❌ 文件未找到: {e}")
        print("请检查文件路径是否正确")
    except PermissionError:
        print("❌ 文件被占用，请关闭Excel文件后重试")
    except Exception as e:
        print(f"❌ 发生错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # 检查是否安装了chardet库
    try:
        import chardet
    except ImportError:
        print("正在安装chardet库...")
        os.system("pip install chardet")
        import chardet

    update_xop_price()