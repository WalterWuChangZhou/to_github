import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os

# ==================== 配置区 ====================
TARGET_FILE = r"E:\2026年QDII LOF基金估值偏差记录.xlsm"
SHEET_NAME = "162411"
START_ROW = 2

# XOP指数配置
XOP_SOURCE = r"E:\B01Python\PythonProject\A06_xop\xop_daily_nav.csv"
XOP_COLUMN = 7  # G列

# 162411基金净值配置
NAV_SOURCE = r"E:\B01Python\PythonProject\A01_fund_nav_data\fund_nav_data\162411.csv"
NAV_COLUMN = 2  # B列
# ================================================


def read_csv_data(file_path, date_col_name=None, value_col_name=None):
    """读取CSV，返回日期→数值的字典"""
    if not os.path.exists(file_path):
        print(f"   ❌ 文件不存在: {file_path}")
        return None

    df = pd.read_csv(file_path)
    print(f"   读取 {len(df)} 行, 列名: {df.columns.tolist()}")

    # 自动识别日期列
    date_col = None
    for col in df.columns:
        if 'date' in col.lower() or '时间' in col.lower() or '日期' in col.lower():
            date_col = col
            break
    if date_col is None:
        date_col = df.columns[0]  # 默认第一列

    # 自动识别数值列
    value_col = None
    for col in df.columns:
        if 'nav' in col.lower() or '净' in col.lower() or '指数' in col.lower() or 'price' in col.lower():
            value_col = col
            break
    if value_col is None:
        value_col = df.columns[1]  # 默认第二列

    print(f"   日期列: {date_col}, 数值列: {value_col}")

    # 建立映射
    result = {}
    try:
        df[date_col] = pd.to_datetime(df[date_col])
        for _, row in df.iterrows():
            key = row[date_col].strftime('%Y-%m-%d')
            result[key] = float(row[value_col])
    except Exception as e:
        print(f"   ❌ 数据转换失败: {e}")
        return None

    print(f"   ✅ 加载 {len(result)} 条数据")
    return result


def write_to_excel(target_file, sheet_name, data_dict, target_col, start_row=2):
    """将数据写入Excel指定列"""
    # 打开Excel
    wb = load_workbook(target_file, keep_vba=True,data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"   ❌ 找不到工作表: {sheet_name}")
        wb.close()
        return 0

    ws = wb[sheet_name]
    max_row = ws.max_row
    updated = 0
    not_found = []

    for row in range(start_row, max_row + 1):
        date_cell = ws.cell(row, 1)  # A列日期
        if date_cell.value is None:
            continue

        # 统一日期格式
        try:
            if isinstance(date_cell.value, datetime):
                row_date = date_cell.value.strftime('%Y-%m-%d')
            else:
                row_date = pd.to_datetime(date_cell.value).strftime('%Y-%m-%d')
        except:
            continue

        # 匹配写入
        if row_date in data_dict:
            ws.cell(row, target_col, data_dict[row_date])
            updated += 1
        else:
            if len(not_found) < 5:
                not_found.append(row_date)

    # 保存
    wb.save(target_file)
    wb.close()

    if not_found:
        print(f"   ⚠️ 未匹配日期样例: {not_found}")

    return updated


def xop_nav():
    print("=" * 70)
    print("📊 162411页面数据写入工具")
    print("=" * 70)

    # 检查目标文件是否存在
    if not os.path.exists(TARGET_FILE):
        print(f"❌ 目标文件不存在: {TARGET_FILE}")
        return

    # 1. 读取XOP指数数据
    print("\n📂 读取XOP指数数据:")
    xop_data = read_csv_data(XOP_SOURCE)
    if xop_data is None:
        print("   ❌ XOP数据读取失败，退出")
        return

    # 2. 读取162411基金净值数据
    print("\n📂 读取162411基金净值数据:")
    nav_data = read_csv_data(NAV_SOURCE)
    if nav_data is None:
        print("   ❌ 基金净值数据读取失败，退出")
        return

    # 3. 写入XOP指数到G列
    print(f"\n✏️ 写入XOP指数到G列 (第{XOP_COLUMN}列)...")
    xop_count = write_to_excel(TARGET_FILE, SHEET_NAME, xop_data, XOP_COLUMN, START_ROW)
    print(f"   ✅ 更新 {xop_count} 条")

    # 4. 写入基金净值到B列
    print(f"\n✏️ 写入162411净值到B列 (第{NAV_COLUMN}列)...")
    nav_count = write_to_excel(TARGET_FILE, SHEET_NAME, nav_data, NAV_COLUMN, START_ROW)
    print(f"   ✅ 更新 {nav_count} 条")

    # 5. 统计
    print("\n" + "=" * 70)
    print("📊 执行完成！")
    print("=" * 70)
    print(f"   📅 XOP数据范围: {min(xop_data.keys())} 至 {max(xop_data.keys())}")
    print(f"   📅 基金净值范围: {min(nav_data.keys())} 至 {max(nav_data.keys())}")
    print(f"   ✅ XOP写入: {xop_count} 条")
    print(f"   ✅ 基金净值写入: {nav_count} 条")
    print(f"   📂 目标文件: {TARGET_FILE}")
    print(f"   📋 工作表: {SHEET_NAME}")
    print("=" * 70)


if __name__ == "__main__":
    xop_nav()