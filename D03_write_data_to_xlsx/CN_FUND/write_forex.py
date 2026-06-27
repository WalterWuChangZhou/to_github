import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
import shutil
import traceback

# file path with pathlib
source_file = Path(r'E:\B01Python\PythonProject\A05_fx_rate\fx_rate\fx_rate.csv')
target_file = Path(r'e:\2026年LOF基金估值偏差记录.xlsm')
backup_file = target_file.with_name(f"{target_file.stem}_backup{target_file.suffix}")


def parse_date(value):
        """
        将多种格式的输入解析为 datetime.date 对象
        """
        # 1. 已经是 datetime
        if isinstance(value, datetime):
            return value.date()

        # 2. 已经是 date（防止误传）
        if isinstance(value, date):
            return value

        # 3. 处理字符串
        if isinstance(value, str):
            value = value.strip()  # 去除首尾空格
            if not value:  # 空字符串
                return None

            # 常见日期格式
            formats = [
                '%Y-%m-%d', '%Y/%m/%d',
                '%m/%d/%Y', '%d/%m/%Y',
                '%Y%m%d',  # 20260626
                '%b %d, %Y',  # Jun 26, 2026
            ]
            for fmt in formats:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue

            # 尝试 pandas 万能解析
            try:
                return pd.to_datetime(value).date()
            except (ValueError, TypeError):
                return None

        # 4. 处理数字（Excel 日期序列号）
        if isinstance(value, (int, float)):
            try:
                # Excel 1900 日期系统（考虑闰年bug）
                return pd.to_datetime(value, unit='D', origin='1899-12-30').date()
            except:
                return None

        # 5. 最终失败
        return None

def find_columns(ws):
    """查找日期和汇率列"""
    date_col = None
    rate_col = None

    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if not cell_value:
            continue

        col_name = str(cell_value).strip().upper()

        if col_name in ['DATE', '日期', '交易日']:
            date_col = col_idx
            print(f"  ✅ 找到日期列: 第{col_idx}列 '{cell_value}'")

        if 'HKD' in col_name and 'CNY' in col_name:
            rate_col = col_idx
            print(f"  ✅ 找到HKD/CNY列: 第{col_idx}列 '{cell_value}'")

        if date_col and rate_col:
            break

    return date_col, rate_col


def update_usd_cny_safe():
    """安全更新USD/CNY汇率"""
    try:
        # 1. 验证文件
        if not source_file.exists():
            print(f"❌ 源文件不存在: {source_file}")
            return False

        if not target_file.exists():
            print(f"❌ 目标文件不存在: {target_file}")
            return False

        # 2. 备份原文件
        print(f"📦 正在备份原文件到: {backup_file}")
        shutil.copy2(target_file, backup_file)

        # 3. 读取源数据（优化）
        print("📂 正在读取源文件...")
        df_source = pd.read_csv(source_file)
        df_source['Date'] = pd.to_datetime(df_source['Date'])

        # 构建高效的汇率字典
        dates = df_source['Date'].dt.date
        rates = df_source['USD/CNY']
        rate_dict = dict(zip(dates, rates))

        # 添加字符串格式的键
        date_str1 = df_source['Date'].dt.strftime('%Y-%m-%d')
        date_str2 = df_source['Date'].dt.strftime('%Y/%m/%d')
        rate_dict.update(dict(zip(date_str1, rates)))
        rate_dict.update(dict(zip(date_str2, rates)))

        print(f"  ✅ 加载 {len(df_source)} 条汇率数据")
        print(f"  📅 日期范围: {df_source['Date'].min().date()} 至 {df_source['Date'].max().date()}")

        # 4. 打开Excel文件
        print("📂 正在打开目标文件...")
        wb = load_workbook(target_file, keep_vba=True, data_only=False)

        total_updated = 0
        processed_sheets = []

        # 5. 处理每个工作表
        for sheet_name in wb.sheetnames:
            print(f"\n📋 处理工作表: {sheet_name}")
            ws = wb[sheet_name]

            # 查找列
            date_col, rate_col = find_columns(ws)

            if date_col is None:
                print("  ⚠️ 跳过: 未找到日期列")
                continue

            if rate_col is None:
                print("  ⚠️ 跳过: 未找到USD/CNY列")
                continue

            # 更新数据
            updated_count = 0
            total_rows = 0

            for row_idx in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row_idx, column=date_col)
                if date_cell.value is None:
                    continue

                total_rows += 1
                date_key = parse_date(date_cell.value)

                if date_key is None:
                    continue

                # 尝试匹配汇率
                rate = None
                if date_key in rate_dict:
                    rate = rate_dict[date_key]
                else:
                    date_str1 = date_key.strftime('%Y-%m-%d')
                    date_str2 = date_key.strftime('%Y/%m/%d')
                    if date_str1 in rate_dict:
                        rate = rate_dict[date_str1]
                    elif date_str2 in rate_dict:
                        rate = rate_dict[date_str2]

                if rate is not None:
                    ws.cell(row=row_idx, column=rate_col, value=rate)
                    updated_count += 1

                    # 显示前3条示例
                    if updated_count <= 3:
                        print(f"  ✅ 行{row_idx}: {date_key} -> {rate}")

            print(f"  📊 总计 {total_rows} 行日期，成功更新 {updated_count} 行")
            total_updated += updated_count
            processed_sheets.append(sheet_name)

        # 6. 保存文件
        print("\n💾 正在保存文件...")
        wb.save(target_file)
        wb.close()

        # 7. 输出结果
        print("\n" + "=" * 60)
        print("✅ 更新完成！")
        print(f"  📋 处理工作表: {processed_sheets}")
        print(f"  📊 总共更新: {total_updated} 行")
        print(f"  💾 备份文件: {backup_file}")
        print("=" * 60)

        return True

    except PermissionError:
        print("❌ 权限错误！请确保目标文件未被Excel打开")
        print("   请关闭Excel文件后重新运行")
        return False

    except Exception as e:
        print(f"❌ 发生错误: {e}")
        traceback.print_exc()

        # 自动恢复备份
        if backup_file.exists():
            print(f"\n⚠️ 正在从备份恢复...")
            shutil.copy2(backup_file, target_file)
            print(f"✅ 已恢复原文件")

        return False


if __name__ == "__main__":
    update_usd_cny_safe()