import pandas as pd
import os
import shutil
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import time


class LOFDataUpdater:
    def __init__(self, target_file):
        """
        初始化数据更新器
        :param target_file: 目标Excel文件路径
        """
        self.target_file = target_file
        # 修正备份文件路径
        self.backup_file = r"E:\2026年LOF基金估值偏差记录_backup.xlsm"

        # 数据源路径
        self.index_data_path = r"E:\B01Python\PythonProject\A02_index_and_fund_close_data\index_data"
        self.nav_data_path = r"E:\B01Python\PythonProject\A01_fund_nav_data\fund_nav_data"
        self.share_data_path = r"E:\B01Python\PythonProject\A03_fund_share_data\fund_share_data"
        self.fx_rate_path = r"E:\B01Python\PythonProject\A05_fx_rate\fx_rate\fx_rate.csv"

    def backup_target_file(self):
        """备份目标文件"""
        try:
            if os.path.exists(self.target_file):
                # 如果备份文件已存在，先删除
                if os.path.exists(self.backup_file):
                    try:
                        os.remove(self.backup_file)
                        print(f"删除旧备份文件：{self.backup_file}")
                    except Exception as e:
                        print(f"删除旧备份文件失败：{e}")
                        # 如果无法删除，使用带时间戳的新备份文件名
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        self.backup_file = f"E:\\2026年LOF基金估值偏差记录_backup_{timestamp}.xlsm"
                        print(f"使用新备份文件名：{self.backup_file}")

                shutil.copy2(self.target_file, self.backup_file)
                print(f"备份成功：{self.backup_file}")
                return True
            else:
                print(f"目标文件不存在：{self.target_file}")
                return False
        except Exception as e:
            print(f"备份失败：{e}")
            return False

    def try_close_file(self, file_path, max_attempts=3):
        """
        尝试关闭可能占用文件的程序
        :param file_path: 文件路径
        :param max_attempts: 最大尝试次数
        """
        for attempt in range(max_attempts):
            try:
                # 尝试以读写模式打开文件
                with open(file_path, 'r+b') as f:
                    f.close()
                return True
            except PermissionError:
                print(f"文件被占用，等待重试... (尝试 {attempt + 1}/{max_attempts})")
                time.sleep(2)
            except Exception as e:
                print(f"检查文件状态时出错：{e}")
                time.sleep(1)
        return False

    def read_index_data(self, fund_code, index_code):
        """
        读取指数和基金收盘价数据
        :param fund_code: 基金代码
        :param index_code: 指数代码
        :return: DataFrame
        """
        file_pattern = f"{fund_code}_{index_code}"
        file_path = None

        # 查找匹配的文件
        for file in os.listdir(self.index_data_path):
            if file.endswith('.csv') and file.startswith(file_pattern):
                file_path = os.path.join(self.index_data_path, file)
                break

        if file_path and os.path.exists(file_path):
            try:
                df = pd.read_csv(file_path)
                # 确保列名统一为小写
                df.columns = df.columns.str.lower()
                return df
            except Exception as e:
                print(f"读取指数数据失败 {file_path}: {e}")
                return None
        else:
            print(f"未找到指数数据文件：{fund_code}_{index_code}")
            return None

    def read_nav_data(self, fund_code):
        """
        读取净值数据
        :param fund_code: 基金代码
        :return: DataFrame
        """
        file_pattern = f"{fund_code}"
        file_path = None

        for file in os.listdir(self.nav_data_path):
            if file.endswith('.csv') and file.startswith(file_pattern):
                file_path = os.path.join(self.nav_data_path, file)
                break

        if file_path and os.path.exists(file_path):
            try:
                df = pd.read_csv(file_path)
                df.columns = df.columns.str.lower()
                return df
            except Exception as e:
                print(f"读取净值数据失败 {file_path}: {e}")
                return None
        else:
            print(f"未找到净值数据文件：{fund_code}")
            return None

    def read_share_data(self, fund_code):
        """
        读取份额数据
        :param fund_code: 基金代码
        :return: DataFrame
        """
        file_pattern = f"{fund_code}"
        file_path = None

        for file in os.listdir(self.share_data_path):
            if file.endswith('.csv') and file.startswith(file_pattern):
                file_path = os.path.join(self.share_data_path, file)
                break

        if file_path and os.path.exists(file_path):
            try:
                df = pd.read_csv(file_path)
                df.columns = df.columns.str.lower()
                return df
            except Exception as e:
                print(f"读取份额数据失败 {file_path}: {e}")
                return None
        else:
            print(f"未找到份额数据文件：{fund_code}")
            return None

    def read_fx_rate_data(self):
        """读取汇率数据"""
        if os.path.exists(self.fx_rate_path):
            try:
                df = pd.read_csv(self.fx_rate_path)
                df.columns = df.columns.str.lower()

                # 打印列名以便调试
                print(f"汇率数据列名：{df.columns.tolist()}")
                print(f"汇率数据行数：{len(df)}")
                print(f"汇率数据前5行：\n{df.head()}")

                # 确保日期列是datetime类型
                if 'date' in df.columns:
                    df['date'] = pd.to_datetime(df['date']).dt.date
                    print(f"汇率数据日期范围：{df['date'].min()} 到 {df['date'].max()}")
                else:
                    print("汇率数据中没有日期列")

                return df
            except Exception as e:
                print(f"读取汇率数据失败：{e}")
                return None
        else:
            print(f"汇率文件不存在：{self.fx_rate_path}")
            return None

    def update_worksheet_data(self, worksheet, data_df, column_mapping):
        """
        更新工作表数据
        :param worksheet: openpyxl工作表对象
        :param data_df: 数据DataFrame
        :param column_mapping: 列名映射字典 {目标列名: 源列名}
        """
        if data_df is None or data_df.empty:
            print("数据为空，跳过更新")
            return

        # 获取工作表中的日期列数据
        date_col = 'date'
        if date_col not in data_df.columns:
            print(f"数据中缺少日期列：{date_col}")
            return

        # 将日期列统一为date类型
        data_df[date_col] = pd.to_datetime(data_df[date_col]).dt.date

        # 获取工作表的所有行（从第1行开始，第0行是表头）
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        # 获取表头行（第1行）
        header_row = 1
        header_dict = {}
        for col in range(1, max_col + 1):
            cell_value = worksheet.cell(row=header_row, column=col).value
            if cell_value:
                header_dict[str(cell_value).lower()] = col

        # 打印表头以便调试
        print(f"工作表表头：{list(header_dict.keys())}")

        # 找到日期列和需要更新的列
        if 'date' not in header_dict:
            print("工作表中未找到日期列")
            return

        date_col_idx = header_dict['date']

        # 构建需要更新的列索引映射
        update_cols = {}
        for target_col, source_col in column_mapping.items():
            if target_col.lower() in header_dict:
                update_cols[header_dict[target_col.lower()]] = source_col
            else:
                print(f"工作表中未找到列：{target_col}")

        if not update_cols:
            print("没有需要更新的列")
            return

        # 创建日期字典，加快查找速度
        data_date_dict = {}
        for idx, row in data_df.iterrows():
            date_val = row[date_col]
            if isinstance(date_val, (datetime, pd.Timestamp)):
                date_val = date_val.date()
            data_date_dict[date_val] = row

        # 遍历工作表的每一行数据（从第2行开始）
        updated_count = 0
        for row_idx in range(2, max_row + 1):
            date_cell = worksheet.cell(row=row_idx, column=date_col_idx)
            if date_cell.value:
                # 处理日期格式
                current_date = date_cell.value
                if isinstance(current_date, datetime):
                    current_date = current_date.date()
                elif isinstance(current_date, pd.Timestamp):
                    current_date = current_date.date()
                elif isinstance(current_date, str):
                    try:
                        current_date = datetime.strptime(current_date, '%Y-%m-%d').date()
                    except:
                        try:
                            current_date = datetime.strptime(current_date, '%Y/%m/%d').date()
                        except:
                            continue

                # 在数据中查找匹配的日期
                if current_date in data_date_dict:
                    data_row = data_date_dict[current_date]
                    # 更新各个列
                    for col_idx, source_col in update_cols.items():
                        if source_col in data_row:
                            value = data_row[source_col]
                            # 处理NaN值
                            if pd.isna(value):
                                value = None
                            # 如果是数值，转换为适当类型
                            if isinstance(value, (int, float)):
                                worksheet.cell(row=row_idx, column=col_idx, value=float(value))
                            else:
                                worksheet.cell(row=row_idx, column=col_idx, value=value)
                    updated_count += 1

        print(f"更新了 {updated_count} 行数据")

    def update_fx_rate_for_worksheet(self, worksheet, fx_df):
        """
        更新汇率数据到指定工作表（161124）
        :param worksheet: 工作表对象
        :param fx_df: 汇率DataFrame
        """
        if fx_df is None or fx_df.empty:
            print("汇率数据为空")
            return

        # 获取表头
        header_row = 1
        header_dict = {}
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=header_row, column=col).value
            if cell_value:
                header_dict[str(cell_value).lower()] = col

        print(f"161124工作表表头：{list(header_dict.keys())}")

        if 'date' not in header_dict:
            print("工作表中缺少日期列")
            return

        # 检查HKD/CNY列（注意：可能列名是"HKD/CNY"或"hkd/cny"）
        hkd_col_name = None
        for col_name in header_dict.keys():
            if 'hkd' in col_name and 'cny' in col_name:
                hkd_col_name = col_name
                break

        if not hkd_col_name:
            print("工作表中未找到HKD/CNY列")
            print(f"可用的列名：{list(header_dict.keys())}")
            return

        date_col_idx = header_dict['date']
        hkd_col_idx = header_dict[hkd_col_name]

        print(f"找到HKD/CNY列：{hkd_col_name}，列索引：{hkd_col_idx}")

        # 创建汇率日期字典
        fx_dict = {}
        for idx, row in fx_df.iterrows():
            date_val = row['date']
            if isinstance(date_val, (datetime, pd.Timestamp)):
                date_val = date_val.date()
            elif isinstance(date_val, str):
                try:
                    date_val = datetime.strptime(date_val, '%Y-%m-%d').date()
                except:
                    continue

            # 查找HKD/CNY列
            if 'hkd/cny' in row:
                fx_dict[date_val] = row['hkd/cny']
            else:
                print(f"汇率数据中未找到HKD/CNY列，可用列：{fx_df.columns.tolist()}")
                return

        print(f"汇率数据字典有 {len(fx_dict)} 条记录")
        if len(fx_dict) > 0:
            sample_date = list(fx_dict.keys())[0]
            sample_value = fx_dict[sample_date]
            print(f"汇率数据示例：日期 {sample_date} -> HKD/CNY: {sample_value}")

        # 更新数据
        updated_count = 0
        for row_idx in range(2, worksheet.max_row + 1):
            date_cell = worksheet.cell(row=row_idx, column=date_col_idx)
            if date_cell.value:
                current_date = date_cell.value
                if isinstance(current_date, datetime):
                    current_date = current_date.date()
                elif isinstance(current_date, pd.Timestamp):
                    current_date = current_date.date()
                elif isinstance(current_date, str):
                    try:
                        current_date = datetime.strptime(current_date, '%Y-%m-%d').date()
                    except:
                        try:
                            current_date = datetime.strptime(current_date, '%Y/%m/%d').date()
                        except:
                            continue

                if current_date in fx_dict:
                    value = fx_dict[current_date]
                    # 确保是数值类型
                    if pd.isna(value):
                        value = None
                    else:
                        value = float(value)
                    worksheet.cell(row=row_idx, column=hkd_col_idx, value=value)
                    updated_count += 1

                    # 打印前几条更新记录以便调试
                    if updated_count <= 5:
                        print(f"更新行 {row_idx}，日期 {current_date}，HKD/CNY: {value}")

        print(f"更新了 {updated_count} 行汇率数据")

    def process_worksheet(self, worksheet_name, workbook):
        """
        处理单个工作表
        :param worksheet_name: 工作表名称
        :param workbook: 工作簿对象
        """
        print(f"\n处理工作表：{worksheet_name}")

        # 获取工作表
        if worksheet_name not in workbook.sheetnames:
            print(f"工作表不存在：{worksheet_name}")
            return

        worksheet = workbook[worksheet_name]

        # 判断工作表名称是否为基金代码
        fund_code = worksheet_name

        # 尝试提取基金代码（如果是161124这种格式）
        if not fund_code.isdigit():
            print(f"工作表名称不是纯数字，跳过：{fund_code}")
            return

        # 特殊情况：161124需要处理汇率
        if fund_code == '161124':
            print("处理特殊工作表161124的汇率数据")
            fx_df = self.read_fx_rate_data()
            if fx_df is not None:
                self.update_fx_rate_for_worksheet(worksheet, fx_df)
            # 继续处理其他数据

        # 寻找对应的指数数据文件
        index_files = []
        try:
            for file in os.listdir(self.index_data_path):
                if file.endswith('.csv') and file.startswith(fund_code + '_'):
                    index_files.append(file)
        except Exception as e:
            print(f"读取指数数据目录失败：{e}")
            return

        if not index_files:
            print(f"未找到基金代码 {fund_code} 的指数数据文件")
        else:
            # 处理每个指数数据文件
            for index_file in index_files:
                # 提取指数代码
                parts = index_file.replace('.csv', '').split('_')
                if len(parts) >= 2:
                    index_code = parts[1]
                    print(f"  处理指数数据：{index_code}")

                    # 读取数据
                    index_df = self.read_index_data(fund_code, index_code)
                    if index_df is not None:
                        # 更新指数收盘价、基金收盘价、基金金额等列
                        column_mapping = {
                            'index_close': 'index_close',
                            'fund_close': 'fund_close',
                            'fund_amount': 'fund_amount'
                        }
                        self.update_worksheet_data(worksheet, index_df, column_mapping)

        # 更新净值数据
        nav_df = self.read_nav_data(fund_code)
        if nav_df is not None:
            print("  更新净值数据")
            column_mapping = {'nav': 'nav'}
            self.update_worksheet_data(worksheet, nav_df, column_mapping)

        # 更新份额数据
        share_df = self.read_share_data(fund_code)
        if share_df is not None:
            print("  更新份额数据")
            column_mapping = {'share': 'share'}
            self.update_worksheet_data(worksheet, share_df, column_mapping)

    def save_workbook_safely(self, workbook, file_path, max_attempts=5):
        """
        安全保存工作簿，包括重试机制
        :param workbook: 工作簿对象
        :param file_path: 保存路径
        :param max_attempts: 最大尝试次数
        """
        for attempt in range(max_attempts):
            try:
                # 保存前先尝试关闭可能打开的文件
                self.try_close_file(file_path, 2)

                # 保存工作簿
                workbook.save(file_path)
                print(f"文件保存成功：{file_path}")
                return True

            except PermissionError as e:
                print(f"保存失败 - 权限错误 (尝试 {attempt + 1}/{max_attempts}): {e}")
                if attempt < max_attempts - 1:
                    print("请关闭可能打开该文件的程序（如Excel）...")
                    time.sleep(3)
                else:
                    print("达到最大尝试次数，保存失败")
                    return False

            except Exception as e:
                print(f"保存失败 (尝试 {attempt + 1}/{max_attempts}): {e}")
                if attempt < max_attempts - 1:
                    time.sleep(2)
                else:
                    print("达到最大尝试次数，保存失败")
                    return False

        return False

    def run(self):
        """执行数据更新"""
        print("开始更新LOF基金数据...")
        print(f"目标文件：{self.target_file}")

        # 1. 备份目标文件
        if not self.backup_target_file():
            print("备份失败，程序终止")
            return

        # 2. 先尝试关闭可能打开的文件
        print("检查文件是否被占用...")
        if not self.try_close_file(self.target_file, 3):
            print("警告：文件可能被其他程序占用，但仍尝试继续...")

        # 3. 加载目标工作簿
        try:
            # 使用openpyxl加载工作簿，保留宏和公式
            workbook = load_workbook(self.target_file, keep_vba=True, data_only=False)
            print(f"成功加载目标文件，包含 {len(workbook.sheetnames)} 个工作表")
        except Exception as e:
            print(f"加载目标文件失败：{e}")
            return

        # 4. 处理每个工作表
        for sheet_name in workbook.sheetnames:
            self.process_worksheet(sheet_name, workbook)

        # 5. 保存工作簿（带重试机制）
        print("\n开始保存文件...")
        save_success = self.save_workbook_safely(workbook, self.target_file)

        if save_success:
            print(f"\n数据更新完成，已保存到：{self.target_file}")
            print(f"备份文件位置：{self.backup_file}")
        else:
            # 如果保存失败，尝试保存到临时文件
            temp_file = f"E:\\temp_update_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsm"
            try:
                workbook.save(temp_file)
                print(f"已保存到临时文件：{temp_file}")
                print("请手动将临时文件内容复制到目标文件")
            except Exception as e:
                print(f"保存到临时文件也失败：{e}")


def main():
    """主函数"""
    target_file = r"E:\2026年LOF基金估值偏差记录.xlsm"

    updater = LOFDataUpdater(target_file)
    updater.run()


if __name__ == "__main__":
    main()