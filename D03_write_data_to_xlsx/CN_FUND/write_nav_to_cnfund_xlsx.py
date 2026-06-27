import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import re
import time


def write_lof_fund_nav():
    """
    写入LOF基金的NAV数据
    源文件：E:\B01Python\PythonProject\A01_fund_nav_data\fund_nav_data\基金代码.csv
    目标文件：2026年QDII LOF基金估值偏差记录.xlsm
    写入位置：对应基金代码页面的NAV列（B列）
    说明：日期列由用户手动维护，脚本只负责写入NAV数据
    """
    print("=" * 50)
    print("3）写入LOF基金的NAV")
    print("=" * 50)

    # 文件路径配置
    source_dir = r"E:\B01Python\PythonProject\A01_fund_nav_data\fund_nav_data"
    target_file = r"e:\2026年LOF基金估值偏差记录.xlsm"

    # 检查源文件夹是否存在
    if not os.path.exists(source_dir):
        print(f"错误：源文件夹不存在 - {source_dir}")
        return

    # 检查目标文件是否存在
    if not os.path.exists(target_file):
        print(f"错误：目标文件不存在 - {target_file}")
        return

    # 检查文件是否被占用
    def is_file_locked(filepath):
        try:
            test_file = open(filepath, 'a')
            test_file.close()
            return False
        except:
            return True

    # 如果文件被占用，等待用户关闭
    if is_file_locked(target_file):
        print(f"\n警告：目标文件正在被其他程序使用！")
        print(f"请关闭 Excel 或其他正在使用该文件的程序。")
        print(f"文件路径: {target_file}")

        input("\n按 Enter 键继续重试...")

        retry_count = 0
        while is_file_locked(target_file) and retry_count < 3:
            print(f"文件仍被占用，等待3秒后重试... (尝试 {retry_count + 1}/3)")
            time.sleep(3)
            retry_count += 1

        if is_file_locked(target_file):
            print("错误：文件仍被占用，请手动关闭后重试。")
            return

    # 获取目标文件的所有页面名称
    try:
        wb_target = load_workbook(target_file, keep_vba=True, data_only=True)
        target_sheets = wb_target.sheetnames
        print(f"目标文件包含 {len(target_sheets)} 个页面: {target_sheets}")
        wb_target.close()
    except Exception as e:
        print(f"读取目标文件失败: {e}")
        return

    # 统计变量
    processed_count = 0
    error_count = 0
    total_update_count = 0
    success_list = []
    error_list = []

    # 遍历源文件夹中的所有CSV文件
    csv_files = [f for f in os.listdir(source_dir) if f.endswith('.csv')]
    print(f"找到 {len(csv_files)} 个CSV文件")

    for filename in csv_files:
        # 提取基金代码（文件名不含扩展名）
        fund_code = os.path.splitext(filename)[0]

        # 检查基金代码是否在目标文件的页面中
        if fund_code not in target_sheets:
            print(f"  跳过 {fund_code}: 目标文件中不存在对应的页面")
            continue

        source_file = os.path.join(source_dir, filename)
        print(f"\n{'=' * 50}")
        print(f"处理基金: {fund_code}")
        print(f"{'=' * 50}")

        try:
            # 读取CSV文件
            df_source = None
            encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig', 'ansi']

            for encoding in encodings:
                try:
                    df_source = pd.read_csv(source_file, encoding=encoding)
                    print(f"  使用编码: {encoding}")
                    break
                except:
                    continue

            if df_source is None:
                print(f"  ✗ 无法读取CSV文件")
                error_count += 1
                error_list.append(f"{fund_code}: 无法读取文件")
                continue

            if df_source.empty:
                print(f"  ✗ 源文件为空")
                error_count += 1
                error_list.append(f"{fund_code}: 文件为空")
                continue

            print(f"  源数据列: {list(df_source.columns)}")
            print(f"  数据行数: {len(df_source)}")

            # 识别日期列和NAV列
            date_col = None
            nav_col = None

            # 日期列识别
            date_patterns = ['日期', '交易日期', 'Date', 'date', '净值日期', '估值日期']
            for col in df_source.columns:
                col_str = str(col).strip()
                if any(pattern in col_str for pattern in date_patterns):
                    date_col = col
                    break

            # 如果没找到日期列，尝试自动识别
            if date_col is None:
                for col in df_source.columns:
                    try:
                        sample = df_source[col].dropna().iloc[:5]
                        for val in sample:
                            if isinstance(val, str):
                                if re.match(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', val):
                                    date_col = col
                                    break
                        if date_col:
                            break
                    except:
                        continue

            # NAV列识别
            nav_patterns = ['NAV', 'nav', '单位净值', '净值', '基金净值']
            for col in df_source.columns:
                col_str = str(col).strip()
                if any(pattern in col_str for pattern in nav_patterns):
                    nav_col = col
                    break

            # 如果没找到NAV列，尝试寻找数值列
            if nav_col is None:
                for col in df_source.columns:
                    if col == date_col:
                        continue
                    try:
                        sample = df_source[col].dropna().iloc[:10]
                        numeric_count = sum(1 for val in sample if isinstance(val, (int, float)) or
                                            (isinstance(val, str) and val.replace('.', '').replace('-', '').isdigit()))
                        if numeric_count >= len(sample) * 0.5:
                            nav_col = col
                            break
                    except:
                        continue

            if date_col is None:
                print(f"  ✗ 无法识别日期列")
                print(f"  可用列: {list(df_source.columns)}")
                error_count += 1
                error_list.append(f"{fund_code}: 无法识别日期列")
                continue

            if nav_col is None:
                print(f"  ✗ 无法识别NAV列")
                print(f"  可用列: {list(df_source.columns)}")
                error_count += 1
                error_list.append(f"{fund_code}: 无法识别NAV列")
                continue

            print(f"  识别日期列: {date_col}")
            print(f"  识别NAV列: {nav_col}")

            # 提取数据
            df_data = df_source[[date_col, nav_col]].copy()
            df_data.columns = ['Date', 'NAV']

            # 解析日期
            def parse_date(date_val):
                if pd.isna(date_val):
                    return None
                if isinstance(date_val, (datetime, pd.Timestamp)):
                    return date_val.strftime('%Y-%m-%d')
                if isinstance(date_val, str):
                    date_str = date_val.strip().strip('"\'')
                    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y']:
                        try:
                            dt = datetime.strptime(date_str, fmt)
                            return dt.strftime('%Y-%m-%d')
                        except:
                            continue
                    try:
                        dt = pd.to_datetime(date_str)
                        return dt.strftime('%Y-%m-%d')
                    except:
                        return None
                return None

            df_data['Date'] = df_data['Date'].apply(parse_date)

            # 解析NAV
            def parse_nav(val):
                if pd.isna(val):
                    return None
                if isinstance(val, (int, float)):
                    return float(val)
                if isinstance(val, str):
                    val_str = val.strip().strip('"\'')
                    val_str = val_str.replace(',', '')
                    try:
                        return float(val_str)
                    except:
                        return None
                return None

            df_data['NAV'] = df_data['NAV'].apply(parse_nav)

            # 删除空值
            df_data = df_data.dropna(subset=['Date', 'NAV'])

            if df_data.empty:
                print(f"  ✗ 没有有效的NAV数据")
                error_count += 1
                error_list.append(f"{fund_code}: 没有有效数据")
                continue

            # 按日期排序
            df_data = df_data.sort_values('Date')

            print(f"  读取到 {len(df_data)} 条NAV数据")
            print(f"  日期范围: {df_data['Date'].min()} 至 {df_data['Date'].max()}")
            print(f"  NAV范围: {df_data['NAV'].min():.4f} 至 {df_data['NAV'].max():.4f}")

            # 打开目标文件
            wb_target = load_workbook(target_file, keep_vba=True, data_only=True)
            ws_target = wb_target[fund_code]

            # 查找NAV列（B列）
            nav_col_target = None
            for col_idx in range(1, ws_target.max_column + 1):
                cell_value = ws_target.cell(row=1, column=col_idx).value
                if cell_value and str(cell_value).strip() == 'NAV':
                    nav_col_target = col_idx
                    break

            if nav_col_target is None:
                nav_col_target = 2  # B列
                print(f"  ⚠ 未找到NAV列，使用B列")

            # 日期列固定为A列
            date_col_target = 1

            print(f"  目标日期列: A列")
            print(f"  目标NAV列: {chr(64 + nav_col_target)}列")

            # 构建日期到行号的映射（从A列读取日期）
            date_row_map = {}
            for row_idx in range(2, ws_target.max_row + 1):
                cell_value = ws_target.cell(row=row_idx, column=date_col_target).value
                if cell_value:
                    if isinstance(cell_value, datetime):
                        date_str = cell_value.strftime('%Y-%m-%d')
                        date_row_map[date_str] = row_idx
                    elif isinstance(cell_value, pd.Timestamp):
                        date_str = cell_value.strftime('%Y-%m-%d')
                        date_row_map[date_str] = row_idx
                    elif isinstance(cell_value, str):
                        try:
                            dt = pd.to_datetime(cell_value.strip())
                            date_str = dt.strftime('%Y-%m-%d')
                            date_row_map[date_str] = row_idx
                        except:
                            continue
                    elif isinstance(cell_value, (int, float)):
                        # Excel日期序列号
                        try:
                            dt = pd.to_datetime(cell_value, unit='D', origin='1899-12-30')
                            date_str = dt.strftime('%Y-%m-%d')
                            date_row_map[date_str] = row_idx
                        except:
                            continue

            print(f"  目标页面有 {len(date_row_map)} 个日期行")

            if len(date_row_map) == 0:
                print(f"  ⚠ 警告：目标页面没有找到日期数据")
                wb_target.close()
                continue

            # 显示日期范围
            target_dates = sorted(date_row_map.keys())
            if target_dates:
                print(f"  目标页面日期范围: {target_dates[0]} 至 {target_dates[-1]}")

            # 更新NAV数据
            update_count = 0
            for _, row in df_data.iterrows():
                date_str = row['Date']
                nav_value = row['NAV']

                if date_str in date_row_map:
                    row_idx = date_row_map[date_str]
                    ws_target.cell(row=row_idx, column=nav_col_target, value=nav_value)
                    update_count += 1

            print(f"  成功更新 {update_count} 条NAV记录")

            # 保存文件
            try:
                wb_target.save(target_file)
                print(f"  ✓ 成功保存文件")
            except PermissionError as e:
                print(f"  ✗ 保存失败：文件被占用")
                print(f"  请关闭Excel文件后重试")
                wb_target.close()
                error_count += 1
                error_list.append(f"{fund_code}: 文件被占用")
                continue

            wb_target.close()

            processed_count += 1
            total_update_count += update_count
            success_list.append(f"{fund_code}: 更新了 {update_count} 条NAV记录")
            print(f"\n  ✓ 成功处理 {fund_code}，更新了 {update_count} 条NAV记录")

        except Exception as e:
            error_count += 1
            error_msg = f"{fund_code}: {str(e)}"
            error_list.append(error_msg)
            print(f"  ✗ 处理失败: {e}")
            import traceback
            traceback.print_exc()

    # 输出总结
    print("\n" + "=" * 50)
    print("写入LOF基金NAV完成！")
    print(f"成功处理: {processed_count} 个基金")
    print(f"失败: {error_count} 个基金")
    print(f"总共更新: {total_update_count} 条记录")

    if success_list:
        print("\n成功列表:")
        for msg in success_list:
            print(f"  {msg}")

    if error_list:
        print("\n错误列表:")
        for msg in error_list:
            print(f"  {msg}")

    return processed_count, error_count, total_update_count


if __name__ == "__main__":
    write_lof_fund_nav()