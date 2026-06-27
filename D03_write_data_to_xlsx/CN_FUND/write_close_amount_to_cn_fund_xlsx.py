import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import re


def process_fund_close_amount_data():
    """
    从源文件夹读取基金收盘和成交额数据（CSV格式），写入目标Excel文件的对应页面
    源文件命名格式：基金代码_指数代码.csv
    需要写入：fund_close 和 fund_amount
    注意：抓取的是fund_close和fund_amount，而不是index_close
    """
    # 文件路径配置
    source_dir = r"E:\B01Python\PythonProject\A02_index_and_fund_close_data\index_data"
    target_file = r"E:\2026年LOF基金估值偏差记录.xlsm"

    # 检查源文件夹是否存在
    if not os.path.exists(source_dir):
        print(f"错误：源文件夹不存在 - {source_dir}")
        return

    # 检查目标文件是否存在
    if not os.path.exists(target_file):
        print(f"错误：目标文件不存在 - {target_file}")
        return

    # 获取目标文件的所有页面名称
    try:
        wb_target = load_workbook(target_file, keep_vba=True, data_only=True)
        target_sheets = wb_target.sheetnames
        print(f"目标文件包含以下页面: {target_sheets}")
        wb_target.close()
    except Exception as e:
        print(f"读取目标文件失败: {e}")
        return

    # 遍历源文件夹中的所有文件
    processed_count = 0
    error_count = 0
    success_list = []
    error_list = []

    for filename in os.listdir(source_dir):
        # 检查文件扩展名（支持.csv）
        if not filename.endswith('.csv'):
            continue

        # 提取基金代码（文件名中下划线前的部分）
        # 例如：162411_XOP.csv -> 162411
        fund_code = filename.split('_')[0] if '_' in filename else os.path.splitext(filename)[0]

        # 检查基金代码是否在目标文件的页面中
        if fund_code not in target_sheets:
            print(f"跳过 {filename}: 基金代码 {fund_code} 在目标文件中不存在对应的页面")
            continue

        source_file = os.path.join(source_dir, filename)
        print(f"\n处理文件: {filename}")
        print(f"基金代码: {fund_code}")
        print(f"源文件: {source_file}")

        try:
            # 读取CSV文件
            # 尝试不同的编码格式
            encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig', 'ansi']
            df_source = None

            for encoding in encodings:
                try:
                    df_source = pd.read_csv(source_file, encoding=encoding)
                    print(f"  使用编码: {encoding}")
                    break
                except UnicodeDecodeError:
                    continue
                except Exception as e:
                    continue

            if df_source is None:
                print(f"  无法读取CSV文件，尝试了所有编码格式")
                continue

            if df_source.empty:
                print(f"  源文件为空")
                continue

            print(f"  数据列: {list(df_source.columns)}")
            print(f"  数据行数: {len(df_source)}")

            # 识别日期列、fund_close列和fund_amount列
            date_col = None
            close_col = None
            amount_col = None

            # 常见列名模式
            date_patterns = ['日期', '交易日期', 'Date', 'date', '时间', 'Time', 'time']
            # fund_close的列名模式（优先匹配fund_close）
            close_patterns = ['fund_close', 'Fund_Close', 'FUND_CLOSE', 'fund_close', 'close'
                            ]
            # fund_amount的列名模式（优先匹配fund_amount）
            amount_patterns = ['amount', 'Fund_Amount', 'FUND_AMOUNT', 'fund amount', 'Fund Amount',
                             '基金成交额', '成交额', '交易金额', 'Amount',  'AMOUNT']

            # 查找日期列
            for col in df_source.columns:
                col_str = str(col).strip()
                if any(pattern in col_str for pattern in date_patterns):
                    date_col = col
                    break

            # 如果没找到日期列，检查列是否包含日期数据
            if date_col is None:
                for col in df_source.columns:
                    try:
                        sample = df_source[col].dropna().iloc[:5]
                        for val in sample:
                            if isinstance(val, str):
                                if re.match(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', val) or \
                                        re.match(r'\d{1,2}[-/]\d{1,2}[-/]\d{4}', val):
                                    date_col = col
                                    break
                        if date_col:
                            break
                    except:
                        continue

            # 查找fund_close列（优先匹配fund_close相关）
            for col in df_source.columns:
                col_str = str(col).strip()
                if any(pattern in col_str for pattern in close_patterns):
                    close_col = col
                    break

            # 如果没找到fund_close，尝试寻找包含close或净值的列
            if close_col is None:
                for col in df_source.columns:
                    if col == date_col:
                        continue
                    col_str = str(col).strip()
                    if any(keyword in col_str for keyword in ['close', 'Close', 'CLOSE', '收盘', '净值', 'NAV', 'nav']):
                        try:
                            sample = df_source[col].dropna().iloc[:10]
                            numeric_count = sum(1 for val in sample if isinstance(val, (int, float)) or
                                                (isinstance(val, str) and val.replace('.', '').replace('-', '').replace(',', '').isdigit()))
                            if numeric_count >= len(sample) * 0.5:
                                close_col = col
                                break
                        except:
                            continue

            # 查找fund_amount列（优先匹配fund_amount相关）
            for col in df_source.columns:
                col_str = str(col).strip()
                if any(pattern in col_str for pattern in amount_patterns):
                    amount_col = col
                    break

            # 如果没找到fund_amount，尝试寻找包含amount或成交额的列
            if amount_col is None:
                for col in df_source.columns:
                    if col == date_col or col == close_col:
                        continue
                    col_str = str(col).strip()
                    if any(keyword in col_str for keyword in ['amount', 'Amount', 'AMOUNT', '成交额', '交易额', '金额']):
                        try:
                            sample = df_source[col].dropna().iloc[:10]
                            numeric_count = sum(1 for val in sample if isinstance(val, (int, float)) or
                                                (isinstance(val, str) and val.replace('.', '').replace('-', '').replace(',', '').isdigit()))
                            if numeric_count >= len(sample) * 0.5:
                                amount_col = col
                                break
                        except:
                            continue

            # 如果还没找到，尝试找任意数值列作为备选
            if close_col is None:
                for col in df_source.columns:
                    if col == date_col or col == amount_col:
                        continue
                    try:
                        sample = df_source[col].dropna().iloc[:10]
                        numeric_count = sum(1 for val in sample if isinstance(val, (int, float)) or
                                            (isinstance(val, str) and val.replace('.', '').replace('-', '').replace(',', '').isdigit()))
                        if numeric_count >= len(sample) * 0.5:
                            close_col = col
                            print(f"  找到数值列作为close备选: {close_col}")
                            break
                    except:
                        continue

            if amount_col is None:
                for col in df_source.columns:
                    if col == date_col or col == close_col:
                        continue
                    try:
                        sample = df_source[col].dropna().iloc[:10]
                        numeric_count = sum(1 for val in sample if isinstance(val, (int, float)) or
                                            (isinstance(val, str) and val.replace('.', '').replace('-', '').replace(',', '').isdigit()))
                        if numeric_count >= len(sample) * 0.5:
                            amount_col = col
                            print(f"  找到数值列作为amount备选: {amount_col}")
                            break
                    except:
                        continue

            if date_col is None:
                print(f"  无法识别日期列")
                print(f"  可用列: {list(df_source.columns)}")
                continue

            print(f"  识别日期列: {date_col}")
            if close_col:
                print(f"  识别fund_close列: {close_col}")
            else:
                print(f"  警告: 未找到fund_close列")

            if amount_col:
                print(f"  识别fund_amount列: {amount_col}")
            else:
                print(f"  警告: 未找到fund_amount列")

            # 准备数据（提取日期、fund_close和fund_amount）
            columns_to_extract = [date_col]
            if close_col:
                columns_to_extract.append(close_col)
            if amount_col:
                columns_to_extract.append(amount_col)

            df_data = df_source[columns_to_extract].copy()

            # 重命名列
            rename_map = {date_col: 'Date'}
            if close_col:
                rename_map[close_col] = 'fund_close'
            if amount_col:
                rename_map[amount_col] = 'fund_amount'

            df_data = df_data.rename(columns=rename_map)

            # 处理日期格式
            def parse_date(date_val):
                if pd.isna(date_val):
                    return None
                if isinstance(date_val, (datetime, pd.Timestamp)):
                    return date_val.strftime('%Y-%m-%d')
                if isinstance(date_val, str):
                    date_str = date_val.strip().strip('"\'')
                    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y',
                                '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S']:
                        try:
                            dt = datetime.strptime(date_str, fmt)
                            return dt.strftime('%Y-%m-%d')
                        except:
                            continue
                    try:
                        dt = pd.to_datetime(date_str)
                        return dt.strftime('%Y-%m-%d')
                    except:
                        return None
                return None

            df_data['Date'] = df_data['Date'].apply(parse_date)

            # 处理数值数据
            def parse_numeric(val):
                if pd.isna(val):
                    return None
                if isinstance(val, (int, float)):
                    return float(val)
                if isinstance(val, str):
                    val_str = val.strip().strip('"\'')
                    val_str = val_str.replace('%', '').replace(',', '')
                    if ',' in val_str:
                        val_str = val_str.replace(',', '')
                    try:
                        return float(val_str)
                    except:
                        return None
                return None

            if 'fund_close' in df_data.columns:
                df_data['fund_close'] = df_data['fund_close'].apply(parse_numeric)

            if 'fund_amount' in df_data.columns:
                df_data['fund_amount'] = df_data['fund_amount'].apply(parse_numeric)

            # 删除日期为空的行
            df_data = df_data.dropna(subset=['Date'])

            if df_data.empty:
                print(f"  没有有效的日期数据")
                continue

            # 按日期排序
            df_data = df_data.sort_values('Date')

            print(f"  读取到 {len(df_data)} 条数据")
            print(f"  日期范围: {df_data['Date'].min()} 至 {df_data['Date'].max()}")

            # 准备写入目标文件
            wb_target = load_workbook(target_file, keep_vba=True, data_only=True)

            # 获取目标页面
            if fund_code in wb_target.sheetnames:
                ws_target = wb_target[fund_code]
            else:
                print(f"  目标页面 {fund_code} 不存在，跳过")
                wb_target.close()
                continue

            # 查找目标页面中的列
            # 先读取第一行获取列名
            header_row = {}
            for col_idx in range(1, ws_target.max_column + 1):
                cell_value = ws_target.cell(row=1, column=col_idx).value
                if cell_value:
                    header_row[col_idx] = str(cell_value).strip()

            # 查找日期列、fund_close列和fund_amount列
            date_col_target = None
            close_col_target = None
            amount_col_target = None

            for col_idx, col_name in header_row.items():
                if any(pattern in col_name for pattern in ['日期', 'Date', 'date', 'DATE']):
                    date_col_target = col_idx
                elif any(pattern in col_name for pattern in ['CLOSE', 'Close', 'close', '收盘', '净值']):
                    close_col_target = col_idx
                elif any(pattern in col_name for pattern in ['Amount', 'amount', 'AMOUNT', '成交额', '交易额']):
                    amount_col_target = col_idx

            # 如果没找到，使用已知的列位置
            if date_col_target is None:
                date_col_target = 1  # A列
                print(f"  未找到日期列，使用默认A列")

            if close_col_target is None:
                # 根据你提供的结构，CLOSE在F列
                close_col_target = 6  # F列
                print(f"  未找到CLOSE列，使用默认F列")

            if amount_col_target is None:
                # 根据你提供的结构，Amount在G列
                amount_col_target = 7  # G列
                print(f"  未找到Amount列，使用默认G列")

            print(f"  目标页面日期列: {chr(64 + date_col_target)}列 (第{date_col_target}列)")
            print(f"  目标页面CLOSE列: {chr(64 + close_col_target)}列 (第{close_col_target}列)")
            print(f"  目标页面Amount列: {chr(64 + amount_col_target)}列 (第{amount_col_target}列)")

            # 构建目标页面中的日期-行号映射
            date_row_map = {}
            for row_idx in range(2, ws_target.max_row + 1):
                cell_value = ws_target.cell(row=row_idx, column=date_col_target).value
                if cell_value:
                    if isinstance(cell_value, datetime):
                        date_str = cell_value.strftime('%Y-%m-%d')
                    elif isinstance(cell_value, pd.Timestamp):
                        date_str = cell_value.strftime('%Y-%m-%d')
                    elif isinstance(cell_value, str):
                        try:
                            cell_value = cell_value.strip()
                            if cell_value.isdigit() and int(cell_value) > 40000:
                                try:
                                    dt = pd.to_datetime(int(cell_value), unit='D', origin='1899-12-30')
                                    date_str = dt.strftime('%Y-%m-%d')
                                except:
                                    continue
                            else:
                                dt = pd.to_datetime(cell_value)
                                date_str = dt.strftime('%Y-%m-%d')
                        except:
                            continue
                    else:
                        continue
                    date_row_map[date_str] = row_idx

            print(f"  目标页面中有 {len(date_row_map)} 个日期记录")

            # 更新数据（更新fund_close和fund_amount）
            update_close_count = 0
            update_amount_count = 0

            for _, row in df_data.iterrows():
                date_str = row['Date']

                if date_str in date_row_map:
                    row_idx = date_row_map[date_str]

                    # 写入fund_close
                    if 'fund_close' in row and pd.notna(row['fund_close']):
                        ws_target.cell(row=row_idx, column=close_col_target, value=row['fund_close'])
                        update_close_count += 1

                    # 写入fund_amount
                    if 'fund_amount' in row and pd.notna(row['fund_amount']):
                        ws_target.cell(row=row_idx, column=amount_col_target, value=row['fund_amount'])
                        update_amount_count += 1

            # 保存目标文件
            wb_target.save(target_file)
            wb_target.close()

            processed_count += 1
            success_list.append(
                f"{fund_code}: 更新了 {update_close_count} 条close数据，{update_amount_count} 条amount数据")
            print(
                f"  ✓ 成功更新 {fund_code}，更新了 {update_close_count} 条close记录，{update_amount_count} 条amount记录")

        except Exception as e:
            error_count += 1
            error_msg = f"{fund_code} ({filename}): {str(e)}"
            error_list.append(error_msg)
            print(f"  ✗ 处理 {filename} 时出错: {e}")
            import traceback
            traceback.print_exc()

    # 输出处理结果总结
    print("\n" + "=" * 50)
    print("处理完成！")
    print(f"成功处理: {processed_count} 个文件")
    print(f"失败: {error_count} 个文件")

    if success_list:
        print("\n成功列表:")
        for msg in success_list:
            print(f"  {msg}")

    if error_list:
        print("\n错误列表:")
        for msg in error_list:
            print(f"  {msg}")


if __name__ == "__main__":
    process_fund_close_amount_data()