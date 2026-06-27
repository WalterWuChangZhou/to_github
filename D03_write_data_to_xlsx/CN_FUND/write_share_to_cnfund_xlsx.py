import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import re


def process_fund_share_data():
    """
    从源文件夹读取基金份额数据（CSV格式），写入目标Excel文件的对应页面
    源文件命名格式：基金代码.csv
    需要写入：share 数据到目标文件的E列（share列）
    """
    # 文件路径配置
    source_dir = r"E:\B01Python\PythonProject\A03_fund_share_data\fund_share_data"
    target_file = r"e:\2026年LOF基金估值偏差记录.xlsm"

    # 检查源文件夹是否存在
    if not os.path.exists(source_dir):
        print(f"错误：源文件夹不存在 - {source_dir}")
        return

    # 检查目标文件是否存在
    if not os.path.exists(target_file):
        print(f"错误：目标文件不存在 - {target_file}")
        return

    # 获取目标文件的所有页面名称
    try:
        wb_target = load_workbook(target_file, keep_vba=True, data_only=True)
        target_sheets = wb_target.sheetnames
        print(f"目标文件包含以下页面: {target_sheets}")
        wb_target.close()
    except Exception as e:
        print(f"读取目标文件失败: {e}")
        return

    # 遍历源文件夹中的所有文件
    processed_count = 0
    error_count = 0
    success_list = []
    error_list = []

    for filename in os.listdir(source_dir):
        # 检查文件扩展名（支持.csv）
        if not filename.endswith('.csv'):
            continue

        # 提取基金代码（文件名不含扩展名）
        fund_code = os.path.splitext(filename)[0]

        # 检查基金代码是否在目标文件的页面中
        if fund_code not in target_sheets:
            print(f"跳过 {filename}: 基金代码 {fund_code} 在目标文件中不存在对应的页面")
            continue

        source_file = os.path.join(source_dir, filename)
        print(f"\n处理文件: {filename}")
        print(f"基金代码: {fund_code}")
        print(f"源文件: {source_file}")

        try:
            # 读取CSV文件
            # 尝试不同的编码格式
            encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig', 'ansi']
            df_source = None

            for encoding in encodings:
                try:
                    df_source = pd.read_csv(source_file, encoding=encoding)
                    print(f"  使用编码: {encoding}")
                    break
                except UnicodeDecodeError:
                    continue
                except Exception as e:
                    continue

            if df_source is None:
                print(f"  无法读取CSV文件，尝试了所有编码格式")
                continue

            if df_source.empty:
                print(f"  源文件为空")
                continue

            print(f"  数据列: {list(df_source.columns)}")
            print(f"  数据行数: {len(df_source)}")

            # 识别日期列和share列
            date_col = None
            share_col = None

            # 常见列名模式
            date_patterns = ['日期', '交易日期', 'Date', 'date', '时间', 'Time', 'time', '净值日期', '估值日期']
            share_patterns = ['份额', 'share', 'Share', 'SHARE', '持有份额', '基金份额', '总份额', '份额数']

            # 查找日期列
            for col in df_source.columns:
                col_str = str(col).strip()
                if any(pattern in col_str for pattern in date_patterns):
                    date_col = col
                    break

            # 如果没找到日期列，检查列是否包含日期数据
            if date_col is None:
                for col in df_source.columns:
                    try:
                        sample = df_source[col].dropna().iloc[:5]
                        for val in sample:
                            if isinstance(val, str):
                                if re.match(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', val) or \
                                        re.match(r'\d{1,2}[-/]\d{1,2}[-/]\d{4}', val):
                                    date_col = col
                                    break
                        if date_col:
                            break
                    except:
                        continue

            # 查找share列
            for col in df_source.columns:
                col_str = str(col).strip()
                if any(pattern in col_str for pattern in share_patterns):
                    share_col = col
                    break

            # 如果没找到share列，尝试寻找数值列作为备选
            if share_col is None:
                for col in df_source.columns:
                    if col == date_col:
                        continue
                    try:
                        sample = df_source[col].dropna().iloc[:10]
                        numeric_count = sum(1 for val in sample if isinstance(val, (int, float)) or
                                            (isinstance(val, str) and val.replace('.', '').replace('-', '').isdigit()))
                        if numeric_count >= len(sample) * 0.5:
                            share_col = col
                            break
                    except:
                        continue

            if date_col is None:
                print(f"  无法识别日期列")
                print(f"  可用列: {list(df_source.columns)}")
                continue

            if share_col is None:
                print(f"  无法识别share列")
                print(f"  可用列: {list(df_source.columns)}")
                continue

            print(f"  识别日期列: {date_col}")
            print(f"  识别share列: {share_col}")

            # 提取日期和份额数据
            df_data = df_source[[date_col, share_col]].copy()
            df_data.columns = ['Date', 'share']

            # 处理日期格式
            def parse_date(date_val):
                if pd.isna(date_val):
                    return None
                if isinstance(date_val, (datetime, pd.Timestamp)):
                    return date_val.strftime('%Y-%m-%d')
                if isinstance(date_val, str):
                    date_str = date_val.strip().strip('"\'')
                    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y',
                                '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S']:
                        try:
                            dt = datetime.strptime(date_str, fmt)
                            return dt.strftime('%Y-%m-%d')
                        except:
                            continue
                    try:
                        dt = pd.to_datetime(date_str)
                        return dt.strftime('%Y-%m-%d')
                    except:
                        return None
                return None

            df_data['Date'] = df_data['Date'].apply(parse_date)

            # 处理份额数据（转换为数值）
            def parse_share(share_val):
                if pd.isna(share_val):
                    return None
                if isinstance(share_val, (int, float)):
                    return float(share_val)
                if isinstance(share_val, str):
                    share_str = share_val.strip().strip('"\'')
                    # 移除可能的百分号、逗号等
                    share_str = share_str.replace('%', '').replace(',', '')
                    # 处理可能的千分位分隔符
                    if ',' in share_str:
                        share_str = share_str.replace(',', '')
                    try:
                        return float(share_str)
                    except:
                        return None
                return None

            df_data['share'] = df_data['share'].apply(parse_share)

            # 删除空值
            df_data = df_data.dropna(subset=['Date', 'share'])

            if df_data.empty:
                print(f"  没有有效的份额数据")
                continue

            # 按日期排序
            df_data = df_data.sort_values('Date')

            print(f"  读取到 {len(df_data)} 条份额数据")
            print(f"  日期范围: {df_data['Date'].min()} 至 {df_data['Date'].max()}")
            print(f"  份额范围: {df_data['share'].min():.2f} 至 {df_data['share'].max():.2f}")

            # 准备写入目标文件
            wb_target = load_workbook(target_file, keep_vba=True, data_only=True)

            # 获取目标页面
            if fund_code in wb_target.sheetnames:
                ws_target = wb_target[fund_code]
            else:
                print(f"  目标页面 {fund_code} 不存在，跳过")
                wb_target.close()
                continue

            # 设置列位置：A列=DATE，E列=share
            date_col_target = 1  # A列
            share_col_target = 5  # E列

            print(f"  目标页面日期列: A列 (第1列)")
            print(f"  目标页面share列: E列 (第5列)")

            # 验证列名
            date_header = ws_target.cell(row=1, column=date_col_target).value
            share_header = ws_target.cell(row=1, column=share_col_target).value

            if date_header:
                print(f"  日期列标题: {date_header}")
            else:
                print(f"  警告: A列可能不是日期列")

            if share_header:
                print(f"  share列标题: {share_header}")
            else:
                print(f"  警告: E列可能不是share列")

            # 构建目标页面中的日期-行号映射
            date_row_map = {}
            for row_idx in range(2, ws_target.max_row + 1):
                cell_value = ws_target.cell(row=row_idx, column=date_col_target).value
                if cell_value:
                    if isinstance(cell_value, datetime):
                        date_str = cell_value.strftime('%Y-%m-%d')
                    elif isinstance(cell_value, pd.Timestamp):
                        date_str = cell_value.strftime('%Y-%m-%d')
                    elif isinstance(cell_value, str):
                        try:
                            cell_value = cell_value.strip()
                            # 如果是Excel序列号日期
                            if cell_value.isdigit() and int(cell_value) > 40000:
                                try:
                                    dt = pd.to_datetime(int(cell_value), unit='D', origin='1899-12-30')
                                    date_str = dt.strftime('%Y-%m-%d')
                                except:
                                    continue
                            else:
                                dt = pd.to_datetime(cell_value)
                                date_str = dt.strftime('%Y-%m-%d')
                        except:
                            continue
                    else:
                        continue
                    date_row_map[date_str] = row_idx

            print(f"  目标页面中有 {len(date_row_map)} 个日期记录")

            # 更新share数据到E列
            update_count = 0
            for _, row in df_data.iterrows():
                date_str = row['Date']
                share_value = row['share']

                if date_str in date_row_map:
                    row_idx = date_row_map[date_str]
                    # 写入share数据到E列（第5列）
                    ws_target.cell(row=row_idx, column=share_col_target, value=share_value)
                    update_count += 1

            # 保存目标文件
            wb_target.save(target_file)
            wb_target.close()

            processed_count += 1
            success_list.append(f"{fund_code}: 更新了 {update_count} 条份额数据")
            print(f"  ✓ 成功更新 {fund_code}，更新了 {update_count} 条份额记录")

        except Exception as e:
            error_count += 1
            error_msg = f"{fund_code} ({filename}): {str(e)}"
            error_list.append(error_msg)
            print(f"  ✗ 处理 {filename} 时出错: {e}")
            import traceback
            traceback.print_exc()

    # 输出处理结果总结
    print("\n" + "=" * 50)
    print("处理完成！")
    print(f"成功处理: {processed_count} 个文件")
    print(f"失败: {error_count} 个文件")

    if success_list:
        print("\n成功列表:")
        for msg in success_list:
            print(f"  {msg}")

    if error_list:
        print("\n错误列表:")
        for msg in error_list:
            print(f"  {msg}")


if __name__ == "__main__":
    process_fund_share_data()